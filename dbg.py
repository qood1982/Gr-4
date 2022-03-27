import csv
from ctypes import alignment
import re
from pathlib import Path
from majed_modules.multi import *
import openpyxl
from openpyxl.styles import PatternFill, Font, Color, fills, Alignment
# ---------------------------------------------------------

index = list_all_pages('./Assets/Amjad')
owner = index[1]
pages = index[0]
bsd = basic_salary_dictionary(pages)
# -----------------------------------------------------------


def myfilter():
    filterd_list = []
    for page in pages:
        lines = ''
        for line in page.splitlines():
            if main_rgx_genral.search(line):
                lines += line + '\n'
        filterd_list.append(lines)
    return filterd_list


# -----------------------------------------------------------

filtr = myfilter()
indexer = 5
aux = []
total = 0
wb = openpyxl.Workbook()
ws = wb.active
ws.sheet_view.rightToLeft = True
ws.sheet_view.showRowColHeaders = True

# # .................... CULOMN WIDTH SETUP...................................
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 7
ws.column_dimensions["E"].width = 20
# ws.merge_cells('C1:D1')

# # ............STYLE..................
f_style = Font(name="Chalkboard", color="E5E7E9", size=12, bold=True)
fill1 = PatternFill(patternType='solid', fgColor='34495E')
heade_align = Alignment("center")
doc_align = Alignment("center")
headlines = PatternFill(
    patternType='solid',
    fgColor='34495E',
)
col_style = PatternFill(
    patternType='solid',
    fgColor='34495E',
)
# # .............STYLE..............

doc = ws["A6:E395"]
for d in doc:
    for cl in d:
        cl.alignment = doc_align

head = ws["A4:E5"]
for col in head:
    for h in col:
        h.font = f_style
        h.fill = fill1
        h.alignment = heade_align

# # {'gray0625', 'darkDown', 'darkGray',
# #  'lightGrid', 'darkVertical',
# # 'lightTrellis', 'darkHorizontal',
# # 'darkTrellis', 'lightUp', 'lightVertical',
# #  'darkGrid', 'gray125', 'lightDown', 'lightGray',
# #  'solid', 'lightHorizontal', 'mediumGray', 'darkUp'}
# 'Chalknoard'

# # head.fill = fill

# # -------------------------------------------------
ws.cell(row=2, column=1).value = " : الاســـــم"
ws.cell(row=2, column=3).value = "رقم الســجل المــدني:"
ws['D2'].alignment = heade_align
ws.merge_cells('D2:E2')

# # --------------------------------------------------
period = ws.cell(row=5, column=1)
discripe = ws.cell(row=5, column=2)
losses = ws.cell(row=5, column=3)
wage = ws.cell(row=5, column=4)
ref = ws.cell(row=5, column=5)
period.value = 'الفترة'
discripe.value = "الــوصـــف"
losses.value = "الخصميــات"
wage.value = "البنــد"
ref.value = "تاريخ الصفـحة المرجعيــة"

period.font = f_style
discripe.font = f_style
losses.font = f_style
wage.font = f_style
ref.font = f_style

# # --------------------------------------------------------
for x, val in enumerate(filtr):
    payroll = re.search(r'Payroll for (\d*/ \d{4})', val)
    annual = re.search(r'^3000.+', val, re.MULTILINE)
    modefied_payroll = payroll.group(1).replace(('/'), '')
    modefied_payroll = modefied_payroll.replace((' '), '')
    m = modefied_payroll[0:2]
    y = modefied_payroll[2:]
    m_y = y + m

    print(payroll.group(1) + "-" * 100)
    for i, l in enumerate(val.splitlines()):
        s = l.split()

        #         #Gross Amount For Previous Months -------------------------------------------------------------------------------------------------
        #         # ----------------------------------------------------------------------------------------------------------------------------
        if m_y not in l:  #Gross Amount For Previous Months
            if "201903" not in l:
                if re.search(r'^1110.+-', l):
                    wage10o = s[0]
                    discr10o = u'بدل وردية متغيره١٠٪'
                    amount10o = float(s[-2].replace(',', ''))
                    date10o = s[-1]
                    total += amount10o
                    indexer += 1

                    print(
                        f'{wage10o} {discr10o}                                                      {amount10o}     {date10o}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date10o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr10o
                    ws.cell(row=indexer, column=3).value = amount10o
                    ws.cell(row=indexer, column=4).value = int(wage10o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1115.+-',
                               l):  #Gross Amount For Previous Months
                    wage15o = s[0]
                    discr15o = u'تعويض جدول عمل ٥٪'
                    amount15o = float(s[-2].replace(',', ''))
                    date15o = s[-1]
                    total += amount15o
                    indexer += 1
                    print(
                        f'{wage15o} {discr15o}                                                      {amount15o}     {date15o}  out'
                    )

                    ws.cell(row=indexer, column=1).value = date15o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr15o
                    ws.cell(row=indexer, column=3).value = amount15o
                    ws.cell(row=indexer, column=4).value = int(wage15o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1021.+-',
                               l):  #Gross Amount For Previous Months
                    wage21o = s[0]
                    discr21o = u'بدل نائيه ٢٥٪'
                    amount21o = float(s[-2].replace(',', ''))
                    date21o = s[-1]
                    total += amount21o
                    indexer += 1
                    print(
                        f'{wage21o} {discr21o}                                                      {amount21o}     {date21o}  out'
                    )

                    ws.cell(row=indexer, column=1).value = date21o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr21o
                    ws.cell(row=indexer, column=3).value = amount21o
                    ws.cell(row=indexer, column=4).value = int(wage21o)
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1111', l):  #Gross Amount For Previous Months
                    wage11o = s[0]
                    discr11o = u'بدل وردية متغيره١٠٪'
                    amount11o = float(s[-2].replace(',', ''))
                    date11o = s[-1]
                    aux.append(amount11o)
                    if len(aux) == 2:
                        result11o = round(aux[0] + aux[1], 3)
                        total += result11o
                        indexer += 1
                        print(
                            f'{wage11o} {discr11o}                                                      {result11o}     {date11o}  out'
                        )
                        print(x, i)
                        aux.clear()
                        ws.cell(row=indexer, column=1).value = date11o
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr11o
                        ws.cell(row=indexer, column=3).value = result11o
                        ws.cell(row=indexer, column=4).value = int(wage11o)
                        # ws.cell(row=indexer,
                        # column=5).value = u"خصم من شهر سابق"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1113.+-',
                               l):  #Gross Amount For Previous Months
                    wage13o = s[0]
                    discr13o = u'تعويض جدول عمل ٥٪'
                    amount13o = float(s[-2].replace(',', ''))
                    date13o = s[-1]
                    total += amount13o
                    indexer += 1
                    print(
                        f'{wage13o} {discr13o}                                                          {amount13o}     {date13o}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date13o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr13o
                    ws.cell(row=indexer, column=3).value = amount13o
                    ws.cell(row=indexer, column=4).value = int(wage13o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1315.+-',
                               l):  #Gross Amount For Previous Months
                    wage315o = s[0]
                    disc = u"بدل طبيعة عمل ٥٪"
                    amount315o = float(s[-2].replace(',', ''))
                    date315o = s[-1]
                    total += amount315o
                    indexer += 1
                    print(
                        f'{wage315o} {disc}                                                      {amount315o}     {date315o}  out'
                    )
                    ws.cell(row=indexer, column=1).value = date315o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = disc
                    ws.cell(row=indexer, column=3).value = amount315o
                    ws.cell(row=indexer, column=4).value = int(wage315o)
                    # ws.cell(row=indexer,
                    #         column=5).value = u"خصم من نفس الشهر"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1320.+-',
                               l):  #Gross Amount For Previous Months
                    wage320o = s[0]
                    discr320o = u' بدل طبيعة عمل ٢٠٪'
                    amount320o = float(s[-2].replace(',', ''))
                    date320o = s[-1]
                    total += amount320o
                    indexer += 1
                    print(
                        f'{wage320o} {discr320o}                                                         {amount320o}     {date320o}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date320o
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr320o
                    ws.cell(row=indexer, column=3).value = amount320o
                    ws.cell(row=indexer, column=4).value = int(wage320o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

#Gross Amount For Same Month -------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------------

        elif m_y in l:  #Gross Amount For Same Month

            if annual:
                sannual = annual.group().split()
                if "201903" not in l:
                    if re.search(r'^1110', l):
                        wage10i = s[0]
                        discr = u'بدل وردية متغيره١٠٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif10 = round(bsd[date] * 0.1 - amount, 2)
                        total += -dif10
                        indexer += 1
                        print(
                            f'{wage10i} {discr}                      {str(sannual[-1])}                             {-dif10}     {date}  in'
                        )

                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif10
                        ws.cell(row=indexer, column=4).value = int(wage10i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1115', l):  #Gross Amount For Same Month
                        wage15i = s[0]
                        discr = u'تعويض جدول عمل ٥٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif15 = round(bsd[date] * 0.05 - amount, 2)
                        total += -dif15
                        indexer += 1
                        print(
                            f'{wage15i} {discr}                      {str(sannual[-1])}                             {-dif15}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif15
                        ws.cell(row=indexer, column=4).value = int(wage15i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1021', l):  #Gross Amount For Same Month
                        wage21 = s[0]
                        discr = u'بدل نائيه ٢٥٪'
                        amount21 = float(s[-2].replace(',', ''))
                        date21 = s[-1]
                        dif21 = round(bsd[date21] * 0.25 - amount21, 2)
                        total += -dif21
                        indexer += 1
                        print(
                            f'{wage21} {discr}                                                      {-dif21}     {date21}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date21
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = amount21
                        ws.cell(row=indexer, column=4).value = int(wage21)
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1111', l):  #Gross Amount For Same Month
                        wage11i = s[0]
                        discr = u'بدل وردية متغيره١٠٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif11 = round(bsd[date] * 0.1 - amount, 2)
                        total += -dif11
                        indexer += 1
                        print(
                            f'{wage11i} {discr}                      {str(sannual[-1])}                             {-dif11}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif11
                        ws.cell(row=indexer, column=4).value = int(wage11i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1113', l):  #Gross Amount For Same Month
                        wage13i = s[0]
                        discr = u'تعويض جدول عمل ٥٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif13 = round(bsd[date] * 0.05 - amount)
                        total += -dif13
                        indexer += 1
                        print(
                            f'{wage13i} {discr}                      {str(sannual[-1])}                             {-dif13}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif13
                        ws.cell(row=indexer, column=4).value = int(wage13i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1315', l):  #Gross Amount For Same Month
                        wage315i = s[0]
                        disc = u"بدل طبيعة عمل ١٥٪"
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        bsd315 = bsd.get(date) * 0.15
                        dif315 = round(amount - bsd315, 2)
                        total += dif315
                        indexer += 1
                        print(
                            f'{wage315i} {disc}                           {str(sannual[-1])}                           {dif315}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = disc
                        ws.cell(row=indexer, column=3).value = -dif315
                        ws.cell(row=indexer, column=4).value = int(wage315i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1320', l):  #Gross Amount For Same Month
                        wage320i = s[0]
                        discr = u'بدل طبيعة عمل '
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif320 = round(bsd[date] * 0.2 - amount, 2)
                        total += -dif320
                        indexer += 1
                        print(
                            f'{wage320i} {discr}                      {str(sannual[-1])}                             {-dif320}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif320
                        ws.cell(row=indexer, column=4).value = int(wage320i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

# # highlines = ws["A6:E6"]

# column_index = ws[f'A6:A{ws.max_row}']

# for c in column_index:
#     for b in c:
#         b.fill = fill1
#         b.font = f_style
# # column_index.fill = col_style
# # for col in range(1):
# #     # print(col)
# #     for i in range(5, ws.max_row):
# #         ws.cell(row=i, column=1).fill = highlines

ws.cell(row=indexer + 1, column=1).value = 'المـجــمــوع: '
ws.cell(row=indexer + 1, column=3).value = round(total, 2)
ws.cell(row=indexer + 1, column=1).fill = fill1
ws.cell(row=indexer + 1, column=3).fill = fill1
ws.cell(row=indexer + 1, column=1).font = f_style
ws.cell(row=indexer + 1, column=3).font = f_style

ws.merge_cells(f'A{indexer + 1}:B{indexer + 1}')
ws.merge_cells(f'C{indexer + 1}:E{indexer + 1}')
print(f'total is : {total}')
print(ws.max_row)
wb.save(f'./out/{owner}.xlsx')
wb.close()
