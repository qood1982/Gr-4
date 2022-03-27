import csv
from ctypes import alignment
import re
from pathlib import Path
from majed_modules.multi import *
import openpyxl
from openpyxl.styles import PatternFill, Font, Color, fills, Alignment
# ---------------------------------------------------------

index = list_all_pages('./Assets/Qassim Gazwani')
owner = index[1]
pages = index[0]
bsd = basic_salary_dictionary(pages)
# -----------------------------------------------------------


def myfilter():
    filterd_list = []
    for page in pages:
        lines = ''
        for line in page.splitlines():
            if main_rgx_genral.search(line):
                lines += line + '\n'
        filterd_list.append(lines)
    return filterd_list


# -----------------------------------------------------------

filtr = myfilter()
indexer = 5
aux = []
total = 0
wb = openpyxl.Workbook()
ws = wb.active
ws.sheet_view.rightToLeft = True
ws.sheet_view.showRowColHeaders = True

# # .................... CULOMN WIDTH SETUP...................................
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 7
ws.column_dimensions["E"].width = 20
# ws.merge_cells('C1:D1')

# # ............STYLE..................
f_style = Font(name="Chalkboard", color="E5E7E9", size=12, bold=True)
fill1 = PatternFill(patternType='solid', fgColor='34495E')
heade_align = Alignment("center")
doc_align = Alignment("center")
headlines = PatternFill(
    patternType='solid',
    fgColor='34495E',
)
col_style = PatternFill(
    patternType='solid',
    fgColor='34495E',
)
# # .............STYLE..............

doc = ws["A6:E395"]
for d in doc:
    for cl in d:
        cl.alignment = doc_align

head = ws["A4:E5"]
for col in head:
    for h in col:
        h.font = f_style
        h.fill = fill1
        h.alignment = heade_align

# # {'gray0625', 'darkDown', 'darkGray',
# #  'lightGrid', 'darkVertical',
# # 'lightTrellis', 'darkHorizontal',
# # 'darkTrellis', 'lightUp', 'lightVertical',
# #  'darkGrid', 'gray125', 'lightDown', 'lightGray',
# #  'solid', 'lightHorizontal', 'mediumGray', 'darkUp'}
# 'Chalknoard'

# # head.fill = fill

# # -------------------------------------------------
ws.cell(row=2, column=1).value = " : الاســـــم"
ws.cell(row=2, column=3).value = "رقم الســجل المــدني:"
ws['D2'].alignment = heade_align
ws.merge_cells('D2:E2')

# # --------------------------------------------------
period = ws.cell(row=5, column=1)
discripe = ws.cell(row=5, column=2)
losses = ws.cell(row=5, column=3)
wage = ws.cell(row=5, column=4)
ref = ws.cell(row=5, column=5)
period.value = 'الفترة'
discripe.value = "الــوصـــف"
losses.value = "الخصميــات"
wage.value = "البنــد"
ref.value = "تاريخ الصفـحة المرجعيــة"

period.font = f_style
discripe.font = f_style
losses.font = f_style
wage.font = f_style
ref.font = f_style

# # --------------------------------------------------------
for x, val in enumerate(filtr):
    payroll = re.search(r'Payroll for (\d*/ \d{4})', val)
    annual = re.search(r'^3000.+', val, re.MULTILINE)
    modefied_payroll = payroll.group(1).replace(('/'), '')
    modefied_payroll = modefied_payroll.replace((' '), '')
    m = modefied_payroll[0:2]
    y = modefied_payroll[2:]
    m_y = y + m

    print(payroll.group(1) + "-" * 100)
    for i, l in enumerate(val.splitlines()):
        s = l.split()

        #         #Gross Amount For Previous Months -------------------------------------------------------------------------------------------------
        #         # ----------------------------------------------------------------------------------------------------------------------------
        if m_y not in l:  #Gross Amount For Previous Months
            if "201903" not in l:
                if re.search(r'^1110.+-', l):
                    wage10o = s[0]
                    discr = u'بدل وردية متغيره١٠٪'
                    amount = float(s[-2].replace(',', ''))
                    date = s[-1]
                    total += amount
                    indexer += 1

                    print(
                        f'{wage10o} {discr}                                                      {amount}     {date}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr
                    ws.cell(row=indexer, column=3).value = amount
                    ws.cell(row=indexer, column=4).value = int(wage10o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1115.+-',
                               l):  #Gross Amount For Previous Months
                    wage15o = s[0]
                    discr = u'تعويض جدول عمل ٥٪'
                    amount = float(s[-2].replace(',', ''))
                    date = s[-1]
                    total += amount
                    indexer += 1
                    print(
                        f'{wage15o} {discr}                                                      {amount}     {date}  out'
                    )

                    ws.cell(row=indexer, column=1).value = date
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr
                    ws.cell(row=indexer, column=3).value = amount
                    ws.cell(row=indexer, column=4).value = int(wage15o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1111', l):  #Gross Amount For Previous Months
                    wage11o = s[0]
                    discr = u'بدل وردية متغيره١٠٪'
                    amount = float(s[-2].replace(',', ''))
                    dateo11 = s[-1]
                    aux.append(amount)
                    if len(aux) == 2:
                        result = round(aux[0] + aux[1], 3)
                        total += result
                        indexer += 1
                        print(
                            f'{wage11o} {discr}                                                      {result}     {date}  out'
                        )
                        print(x, i)
                        aux.clear()
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = result
                        ws.cell(row=indexer, column=4).value = int(wage11o)
                        # ws.cell(row=indexer,
                        # column=5).value = u"خصم من شهر سابق"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1113.+-',
                               l):  #Gross Amount For Previous Months
                    wage13o = s[0]
                    discr = u'تعويض جدول عمل ٥٪'
                    amount = float(s[-2].replace(',', ''))
                    date = s[-1]
                    total += amount
                    indexer += 1
                    print(
                        f'{wage13o} {discr}                                                          {amount}     {date}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr
                    ws.cell(row=indexer, column=3).value = amount
                    ws.cell(row=indexer, column=4).value = int(wage13o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

                elif re.search(r'^1320.+-|^1315.+-',
                               l):  #Gross Amount For Previous Months
                    wage320o = s[0]
                    discr = u'بدل طبيعة عمل '
                    amount = float(s[-2].replace(',', ''))
                    date = s[-1]
                    total += amount
                    indexer += 1
                    print(
                        f'{wage320o} {discr}                                                         {amount}     {date}  out'
                    )
                    print(x, i)
                    ws.cell(row=indexer, column=1).value = date
                    ws.cell(row=indexer, column=1).fill = fill1
                    ws.cell(row=indexer, column=1).font = f_style
                    ws.cell(row=indexer, column=2).value = discr
                    ws.cell(row=indexer, column=3).value = amount
                    ws.cell(row=indexer, column=4).value = int(wage320o)
                    # ws.cell(row=indexer, column=5).value = u"خصم من شهر سابق"
                    ws.cell(row=indexer, column=5).value = payroll.group(1)

#Gross Amount For Same Month -------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------------

        elif m_y in l:  #Gross Amount For Same Month

            if annual:
                sannual = annual.group().split()
                if "201903" not in l:
                    if re.search(r'^1110', l):
                        wage10i = s[0]
                        discr = u'بدل وردية متغيره١٠٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif10 = round(bsd[date] * 0.1 - amount, 2)
                        total += -dif10
                        indexer += 1
                        print(
                            f'{wage10i} {discr}                      {str(sannual[-1])}                             {-dif10}     {date}  in'
                        )

                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif10
                        ws.cell(row=indexer, column=4).value = int(wage10i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1115', l):  #Gross Amount For Same Month
                        wage15i = s[0]
                        discr = u'تعويض جدول عمل ٥٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif15 = round(bsd[date] * 0.05 - amount, 2)
                        total += -dif15
                        indexer += 1
                        print(
                            f'{wage15i} {discr}                      {str(sannual[-1])}                             {-dif15}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif15
                        ws.cell(row=indexer, column=4).value = int(wage15i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1111', l):  #Gross Amount For Same Month
                        wage11i = s[0]
                        discr = u'بدل وردية متغيره١٠٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif11 = round(bsd[date] * 0.1 - amount, 2)
                        total += -dif11
                        indexer += 1
                        print(
                            f'{wage11i} {discr}                      {str(sannual[-1])}                             {-dif11}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif11
                        ws.cell(row=indexer, column=4).value = int(wage11i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1113', l):  #Gross Amount For Same Month
                        wage13i = s[0]
                        discr = u'تعويض جدول عمل ٥٪'
                        amount = float(s[-2].replace(',', ''))
                        date = s[-1]
                        dif13 = round(bsd[date] * 0.05 - amount)
                        total += -dif13
                        indexer += 1
                        print(
                            f'{wage13i} {discr}                      {str(sannual[-1])}                             {-dif13}     {date}  in'
                        )
                        ws.cell(row=indexer, column=1).value = date
                        ws.cell(row=indexer, column=1).fill = fill1
                        ws.cell(row=indexer, column=1).font = f_style
                        ws.cell(row=indexer, column=2).value = discr
                        ws.cell(row=indexer, column=3).value = -dif13
                        ws.cell(row=indexer, column=4).value = int(wage13i)
                        # ws.cell(row=indexer,
                        #         column=5).value = u"خصم من نفس الشهر"
                        ws.cell(row=indexer, column=5).value = payroll.group(1)

                    elif re.search(r'^1320|^1315',
                                   l):  #Gross Amount For Same Month
                        if l.startswith('1320'):

                            wage320i = s[0]
                            discr = u'بدل طبيعة عمل '
                            amount = float(s[-2].replace(',', ''))
                            date = s[-1]
                            dif320 = round(bsd[date] * 0.2 - amount, 2)
                            total += -dif320
                            indexer += 1
                            print(
                                f'{wage320i} {discr}                      {str(sannual[-1])}                             {-dif320}     {date}  in'
                            )
                            ws.cell(row=indexer, column=1).value = date
                            ws.cell(row=indexer, column=1).fill = fill1
                            ws.cell(row=indexer, column=1).font = f_style
                            ws.cell(row=indexer, column=2).value = discr
                            ws.cell(row=indexer, column=3).value = -dif320
                            ws.cell(row=indexer,
                                    column=4).value = int(wage320i)
                            # ws.cell(row=indexer,
                            #         column=5).value = u"خصم من نفس الشهر"
                            ws.cell(row=indexer,
                                    column=5).value = payroll.group(1)
                        elif l.startswith('1315'):
                            wage315i = s[0]
                            discr = u'5% بدل طبيعة عمل '
                            amount = float(s[-2].replace(',', ''))
                            date = s[-1]
                            dif315 = round(bsd[date] * 0.15 - amount, 2)
                            total += -dif315
                            indexer += 1
                            print(
                                f'{wage315i} {discr}                      {str(sannual[-1])}                             {-dif315}     {date}  in'
                            )
                            ws.cell(row=indexer, column=1).value = date
                            ws.cell(row=indexer, column=1).fill = fill1
                            ws.cell(row=indexer, column=1).font = f_style
                            ws.cell(row=indexer, column=2).value = discr
                            ws.cell(row=indexer, column=3).value = -dif315
                            ws.cell(row=indexer,
                                    column=4).value = int(wage315i)
                            # ws.cell(row=indexer,
                            #         column=5).value = u"خصم من نفس الشهر"
                            ws.cell(row=indexer,
                                    column=5).value = payroll.group(1)
# # highlines = ws["A6:E6"]

# column_index = ws[f'A6:A{ws.max_row}']

# for c in column_index:
#     for b in c:
#         b.fill = fill1
#         b.font = f_style
# # column_index.fill = col_style
# # for col in range(1):
# #     # print(col)
# #     for i in range(5, ws.max_row):
# #         ws.cell(row=i, column=1).fill = highlines

ws.cell(row=indexer + 1, column=1).value = 'المـجــمــوع: '
ws.cell(row=indexer + 1, column=3).value = round(total, 2)
ws.cell(row=indexer + 1, column=1).fill = fill1
ws.cell(row=indexer + 1, column=3).fill = fill1
ws.cell(row=indexer + 1, column=1).font = f_style
ws.cell(row=indexer + 1, column=3).font = f_style

ws.merge_cells(f'A{indexer + 1}:B{indexer + 1}')
ws.merge_cells(f'C{indexer + 1}:E{indexer + 1}')
print(f'total is : {total}')
print(ws.max_row)
wb.save(f'./out/{owner}.xlsx')
wb.close()
